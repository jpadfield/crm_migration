import mysql.connector
from openpyxl import load_workbook
import random, urllib, json, string, csv
from SPARQLWrapper import SPARQLWrapper, JSON
import time
import numpy as np
from rdflib import Graph, Namespace, Literal, BNode
from rdflib.namespace import RDF, RDFS, NamespaceManager, XSD
import os
from pdb import set_trace as st
import requests
import datetime

RRO = Namespace("https://rdf.ng-london.org.uk/raphael/ontology/")
RRI = Namespace("https://rdf.ng-london.org.uk/raphael/resource/")
CRM = Namespace("http://www.cidoc-crm.org/cidoc-crm/")
NGO = Namespace("https://data.ng-london.org.uk/")
AAT = Namespace("http://vocab.getty.edu/page/aat/")
TGN = Namespace("http://vocab.getty.edu/page/tgn/")
WD = Namespace("http://www.wikidata.org/entity/")
DIG = Namespace("http://www.cidoc-crm.org/crmdig/")
SCI = Namespace("http://www.cidoc-crm.org/crmsci/")
OWL = Namespace("http://www.w3.org/2002/07/owl#")
RDF = Namespace("http://www.w3.org/1999/02/22-rdf-syntax-ns#")
RDFS = Namespace("http://www.w3.org/2000/01/rdf-schema#")

def query_graph(graph,subj,pred,obj):
    for s,p,o in graph.triples((subj,pred,obj)):
        print(s.n3(graph.namespace_manager),p.n3(graph.namespace_manager),o.n3(graph.namespace_manager))

def pretty_print_triples(new_graph):
    for x,y,z in new_graph:
        print(x.n3(new_graph.namespace_manager), y.n3(new_graph.namespace_manager), z.n3(new_graph.namespace_manager))

def query_objects(graph, subj, pred, obj, ugp=True):
        objects_list = []
        for s, p, o in graph.triples((subj, pred, obj)):
            if ugp == "full":
              o = str(get_property(o, True, True))
            elif ugp:
              o = str(get_property(o))
            objects_list.append(o)
        return objects_list

# imaging event plus connections between the sample and image added by JP 13/04/22        
def query_subjects(graph, subj, pred, obj, ugp=True):
        subjects_list = []
        for s, p, o in graph.triples((subj, pred, obj)):
            if ugp == "full":
              s = str(get_property(s, True, True))
            elif ugp == "str":
              s = str(s)
            elif ugp:
              s = str(get_property(s))
            subjects_list.append(s)
        return subjects_list

def sparql_query_pythonic(csv_format=True):
    qres = g.query(
        """
        PREFIX rdfs:<http://www.w3.org/2000/01/rdf-schema#>
        PREFIX owl:<http://www.w3.org/2002/07/owl#>
        PREFIX rdf:<http://www.w3.org/1999/02/22-rdf-syntax-ns#>
        PREFIX crm:<http://www.cidoc-crm.org/rdfs/cidoc_crm_v5.0.2_english_label.rdfs#>
        PREFIX rro:<https://rdf.ng-london.org.uk/raphael/ontology/>
        PREFIX rri:<https://rdf.ng-london.org.uk/raphael/resource/>

        SELECT 
        DISTINCT ?image ?server ?pyramid ?category
        WHERE { ?image rro:RP30.has_pyramid ?pyramid .
                ?image rro:RP243.has_pyramid_server ?server . 
                ?image rro:RP98.is_in_project_category ?category }
        """
    )

    if csv_format == True: 
        new_dataframe = []
        for row in qres:
            triples = '%s| %s| %s |%s' % row
            innerlist = []
            if 1 == 1:
                innerlist = triples.split('|')
            new_dataframe.append(innerlist)
        new_dataframe = [x for x in new_dataframe if x]
        return new_dataframe
    else:
        return qres

def map_property(graph,new_graph, old_property, new_property):
    for x, old_property, z in graph.triples((None, old_property, None)):
        #graph.remove((x,y,z))
        new_graph.add((x, new_property, z))
    return graph, new_graph

def map_class(graph,new_graph,old_class,new_class):
    for x,y,z in graph.subjects((old_class,None,None)):
        #graph.remove((x,y,z))
        new_graph.add((new_class,y,z))

    for x,y,z in graph.objects((None,None,old_class)):
        #graph.remove((x,y,z))
        new_graph.add((x,y,new_class))

    return graph, new_graph

def triples_to_csv(triples, filename):
    fields = ['Subject','Predicate','Object']

    with open('outputs/' + filename + '.csv','w',newline='') as f:
        write = csv.writer(f)

        write.writerow(fields)
        write.writerows(triples)

    print('CSV created!')

def triples_to_tsv(triples, filename):
    fields = ['Subject','Predicate','Object']

    with open('outputs/' + filename + '.tsv','w',newline='') as f:
        write = csv.writer(f, delimiter = '\t')

        write.writerow(fields)
        write.writerows(triples)

    print('TSV created!')

def connect_to_sql():
    mydb = mysql.connector.connect(
        host="round4",
        user="sshoc",
        password="IqazZuKaXMmSEqUl"
    )

    return mydb

def get_property(uri, keep_underscores=False, keep_periods=False):
    remove_uri = uri.replace('https://rdf.ng-london.org.uk/raphael/resource/','')
    if keep_underscores is False:
        final_property = remove_uri.replace('_',' ')
    else:
        final_property = remove_uri
        
    if keep_periods is False:
      if '.' in final_property:
        final_property = str(final_property.split('.')[1])
        
    if 'RRR' in final_property:
        final_property = final_property.replace('RRR','')
    return final_property

def get_json(url):
    operUrl = urllib.request.urlopen(url)
    if(operUrl.getcode()==200):
       data = operUrl.read()
       json_data = json.loads(data)
    else:
       print("Error receiving data", operUrl.getcode())
    return json_data

def check_pids_csv(input_literal):
    pid_value = 'None'
    if os.path.isfile('outputs/placeholder_pids.csv') == True:
        with open('outputs/placeholder_pids.csv','r') as f:
            reader = csv.DictReader(f, delimiter=',')
            dict_list = []
            for row in reader:
                dict_list.append(row)
            for value in range(len(dict_list)):
                csv_literal = str(dict_list[value]['Literal value'])
                if csv_literal == str(input_literal):
                    pid_value = str(dict_list[value]['Placeholder PID'])
                    break
    return pid_value

def check_db(input_literal, table_name):
    db = connect_to_sql()
    cursor = db.cursor()
    pid_value = None
    input_literal = str(input_literal)
    if table_name == 'temp_pids':
        val = 'pid_value'
        str_input = 'literal_value'
    elif table_name == 'wikidata':
        val = 'wd_value'
        str_input = 'string_literal'

    query = "SELECT " + val + " FROM sshoc." + table_name + " WHERE " + str_input + " = '" + input_literal + "'"
    cursor.execute(query)
    result = cursor.fetchall()
    
    if len(result) > 0:
        pid_value = result[0][0]

    return pid_value

def check_elasticsearch(search_term, index_name): 
    es = Elasticsearch([{'host':'localhost','port':9200}]) 
    search_term = str(search_term)

    pid_value = None
    body_query = {'query':{'match':{'column1':{'query':search_term, 'fuzziness':0}}}}
    res = es.search(index=index_name,body=body_query)

    for hit in res['hits']['hits']:
        pid_value = hit['_source']['column2']
    
    return pid_value

def find_old_pid(ng_number):
    #json_data = get_json('https://scientific.ng-london.org.uk/export/ngpidexport_00A.json')
    if ng_number.startswith('https'):
        ng_number = str(ng_number.rsplit('/',1)[-1])
    else:
        ng_number = ng_number.replace(' ','_')
    old_pid = None
    
    try:
        export_url = 'https://collectiondata.ng-london.org.uk/es/ng-public/_search?q=identifier.object_number:' + ng_number
    except:
        export_url is None
        
    if export_url is not None:
        data = requests.get(export_url, verify=False)
        json = data.json()
        if 'error' not in json:
            if json['hits']['total'] > 0:
                try:
                    json_ng_number = json['hits']['hits'][0]['_source']['identifier'][0]['object_number']
                except:
                    json_ng_number = None
                if json_ng_number == ng_number:
                    old_pid = json['hits']['hits'][0]['_source']['identifier'][1]['pid_tms']

    return old_pid

def generate_placeholder_PID(input_literal):
    db = connect_to_sql()
    N = 4
    placeholder_PID = ""
    for number in range(N):
        res = ''.join(random.choices(string.ascii_uppercase + string.digits, k = N))
        placeholder_PID += str(res)
        placeholder_PID += '-'
    placeholder_PID = placeholder_PID[:-1]
    
    # Added to make them a different structure to real ones.
    placeholder_PID = "SC-" + placeholder_PID
    
    #input_list = [input_literal, placeholder_PID]
    #fields = ['Literal value','Placeholder PID']
    existing_pid = check_db(input_literal, table_name = 'temp_pids')
    old_pid = find_old_pid(input_literal)
    
    if old_pid is not None:
        return old_pid
    elif existing_pid is not None:
        return existing_pid
    else:
        #db = connect_to_sql()
        cursor = db.cursor()
        input_literal = str(input_literal)
        query = 'INSERT INTO sshoc.temp_pids (literal_value, pid_value) VALUES (%s, %s)'
        val = (input_literal, placeholder_PID)
        cursor.execute(query, val)
        db.commit()

        placeholder_PID = str(placeholder_PID)
        return placeholder_PID

# str() added to subj to avoid errors when combining URLs to strings
def create_PID_from_triple(pid_type, subj):
    if pid_type == 'object':    
        pid_name = str(subj)
    elif pid_type == 'pyramid creation for':
        pid_name = str(pid_type) + ' ' + str(subj)
    else:
        pid_name = str(pid_type) + ' of ' + str(subj)
    #print ("CREATING: " + pid_name)
    output_pid = generate_placeholder_PID(pid_name)
    return output_pid

def find_aat_value(material,material_type):
    material = str(get_property(material))
    if material_type == getattr(RRO,'RP20.has_medium'):
        wb = load_workbook(filename = 'inputs/NG_Medium_and_Support_AAT.xlsx', read_only=True)
        ws = wb['Medium Material']
    elif material_type == getattr(RRO,'RP32.has_support'):
        wb = load_workbook(filename = 'inputs/NG_Medium_and_Support_AAT.xlsx', read_only=True)
        ws = wb['Support Materials']
    elif material_type == 'medium type':
        wb = load_workbook(filename = 'inputs/NG_Medium_and_Support_AAT.xlsx', read_only=True)
        ws = wb['Medium Type']
    elif material_type == 'support type':
        wb = load_workbook(filename = 'inputs/NG_Medium_and_Support_AAT.xlsx', read_only=True)
        ws = wb['Support Type']
    elif material_type == getattr(RRO, 'RP215.has_acted_in_the_role_of_an'):
        wb = load_workbook(filename = 'inputs/NG_Roles_AAT.xlsx', read_only=True)
        ws = wb['AAT_Roles']
    for row in ws.iter_rows(values_only=True):
        if material in row:
            aat_number_string = str(row[1])
            aat_number = aat_number_string.replace('aat:','')
            aat_type = row[2]
            return aat_number, aat_type
    wb.close()

def wikidata_query(literal, literal_type):
    sparql = SPARQLWrapper('https://query.wikidata.org/sparql')
    print('Trying a query with input ' + literal)
    remove_uri = literal.replace('https://rdf.ng-london.org.uk/raphael/resource/','')
    literal = remove_uri.replace('_',' ')
    literal = literal.replace('%C3%A9', 'e')
    literal = literal.replace('%C3%A0', 'a')
    
    if literal_type == 'year':
        thing_type = 'Q577'
        if 'RRR' in literal:
            string_literal = str(literal.rsplit(' ')[-1])
        elif 'About' in literal:
            string_literal = str(literal.rsplit('-')[-1])
        else:
            return None
    elif literal_type == 'institution':
        thing_type = 'Q207694'
        string_literal = str(literal)
    elif literal_type == 'location':
        thing_type = 'Q515'
        string_literal = str(literal)
    elif literal_type == 'language':
        thing_type = 'Q34770'
        string_literal = str(literal)
    
    if string_literal is not None:
        result = check_db(string_literal, 'wikidata')
    else:
        result is None
    
    if result is not None:
        result = result.replace('http://www.wikidata.org/entity/','')
        return result
    else:
        query = ('''
        SELECT DISTINCT ?year
        WHERE
        {
        ?year  wdt:P31 wd:''' + thing_type + ''' .
                ?year rdfs:label ?yearLabel .
        FILTER( str(?yearLabel) = "''' + string_literal + '''" ) .
        SERVICE wikibase:label { bd:serviceParam wikibase:language "[AUTO_LANGUAGE],en". }
        }
        ''')
        
        sparql.setQuery(query)
        sparql.setReturnFormat(JSON)

        try:
          #ret = sparql.queryAndConvert()
          ret = sparql.query().convert()          
        except Exception as e:
          print("ERROR: " + string_literal + " not found in wikidata" )
          ret = {}
          ret["results"] = {}
          ret["results"]["bindings"] = []          
          #print(ret)
          
        #
        time.sleep(3)
        
        if ret["results"]["bindings"] != []:
            for result in ret["results"]["bindings"]:
                db = connect_to_sql()
                cursor = db.cursor()
                result = result["year"]["value"]
                query = 'INSERT INTO sshoc.wikidata (string_literal, wd_value) VALUES (%s, %s)'
                val = (string_literal, result)
                cursor.execute(query, val)
                db.commit()
        else:
            result is None
        
        if result is not None:
            result = result.replace('http://www.wikidata.org/entity/','')

        return result

def run_ruby_program(input_string):
    import subprocess

    ruby_var = 'ruby citation_parser.rb \"' + input_string + '\"'
    output = subprocess.Popen(ruby_var, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    out, error = output.communicate()

    try:
        string_output = out.decode("utf-8")
        json_output = json.loads(string_output)
    except:
        return

    return json_output

def create_year_dates(year):
    import datetime
    year = int(year)

    start = datetime.datetime(year, 1, 1)
    end = datetime.datetime(year, 12, 31)

    return start, end

def display_timing (start_time, loop_start_time, str_prefix="\t"):
  current_time = datetime.datetime.now()
  total_diff = (current_time - start_time)
  loop_diff = (current_time - loop_start_time)
  total_time = (total_diff.total_seconds()) / 60
  loop_time = loop_diff.total_seconds()
  print (str_prefix + "Loop time "+ "%.2f" % loop_time +"sec - Total time: "+ "%.2f" % total_time +"mins")
