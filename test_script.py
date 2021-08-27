from rdflib import Graph, Namespace, Literal, BNode
from rdflib.namespace import RDF, RDFS, NamespaceManager, XSD
from rdflib.serializer import Serializer
from SPARQLWrapper import SPARQLWrapper, JSON
import csv, random, string, json, urllib
import os.path
from openpyxl import load_workbook
from elasticsearch import Elasticsearch
from alive_progress import alive_bar
import requests
import time
import mysql.connector
import re

from pdb import set_trace as st
#exit this with c

RRO = Namespace("https://rdf.ng-london.org.uk/raphael/ontology/")
RRI = Namespace("https://rdf.ng-london.org.uk/raphael/resource/")
CRM = Namespace("http://www.cidoc-crm.org/cidoc-crm/")
NGO = Namespace("https://data.ng-london.org.uk/")
AAT = Namespace("http://vocab.getty.edu/page/aat/")
TGN = Namespace("http://vocab.getty.edu/page/tgn/")
WD = Namespace("http://www.wikidata.org/entity/")
DIG = Namespace("http://www.cidoc-crm.org/crmdig/")
SCI = Namespace("http://www.cidoc-crm.org/crmsci/")
OWL = Namespace("http://www.w3.org/2002/07/owl#")

g = Graph()
g.parse("inputs/rrr_i_v0.5.xml", format="xml")
g.bind('rro',RRO)
g.bind('rri',RRI)

new_graph = Graph()
new_graph.namespace_manager.bind('crm',CRM)
new_graph.namespace_manager.bind('ngo',NGO)
new_graph.namespace_manager.bind('aat',AAT)
new_graph.namespace_manager.bind('tgn',TGN)
new_graph.namespace_manager.bind('wd',WD)
new_graph.namespace_manager.bind('rro',RRO)
new_graph.namespace_manager.bind('rri',RRI)
new_graph.namespace_manager.bind('dig', DIG)
new_graph.namespace_manager.bind('sci', SCI)
new_graph.namespace_manager.bind('owl', OWL)

def query_graph(graph,subj,pred,obj):
    for s,p,o in graph.triples((subj,pred,obj)):
        print(s.n3(new_graph.namespace_manager),p.n3(new_graph.namespace_manager),o.n3(new_graph.namespace_manager))

def query_objects(graph, subj, pred, obj):
        objects_list = []
        for s, p, o in graph.triples((subj, pred, obj)):
            o = str(get_property(o))
            objects_list.append(o)
        return objects_list

def sparql_query_pythonic(csv_format=True):
    qres = g.query(
        """
        PREFIX rdfs:<http://www.w3.org/2000/01/rdf-schema#>
        PREFIX owl:<http://www.w3.org/2002/07/owl#>
        PREFIX rdf:<http://www.w3.org/1999/02/22-rdf-syntax-ns#>
        PREFIX crm:<http://www.cidoc-crm.org/rdfs/cidoc_crm_v5.0.2_english_label.rdfs#>
        PREFIX rro:<https://rdf.ng-london.org.uk/raphael/ontology/>
        PREFIX rri:<https://rdf.ng-london.org.uk/raphael/resource/>

        SELECT 
        DISTINCT ?image ?server ?pyramid ?category
        WHERE { ?image rro:RP30.has_pyramid ?pyramid .
                ?image rro:RP243.has_pyramid_server ?server . 
                ?image rro:RP98.is_in_project_category ?category }
        """
    )

    if csv_format == True: 
        new_dataframe = []
        for row in qres:
            triples = '%s| %s| %s |%s' % row
            innerlist = []
            if 1 == 1:
                innerlist = triples.split('|')
            new_dataframe.append(innerlist)
        new_dataframe = [x for x in new_dataframe if x]
        return new_dataframe
    else:
        return qres

def map_property(graph,new_graph, old_property, new_property):
    for x, old_property, z in graph.triples((None, old_property, None)):
        #graph.remove((x,y,z))
        new_graph.add((x, new_property, z))
    return graph, new_graph

def map_class(graph,new_graph,old_class,new_class):
    for x,y,z in graph.subjects((old_class,None,None)):
        #graph.remove((x,y,z))
        new_graph.add((new_class,y,z))

    for x,y,z in graph.objects((None,None,old_class)):
        #graph.remove((x,y,z))
        new_graph.add((x,y,new_class))

    return graph, new_graph

def triples_to_csv(triples, filename):
    fields = ['Subject','Predicate','Object']

    with open('outputs/' + filename + '.csv','w',newline='') as f:
        write = csv.writer(f)

        write.writerow(fields)
        write.writerows(triples)

    print('CSV created!')

def triples_to_tsv(triples, filename):
    fields = ['Subject','Predicate','Object']

    with open('outputs/' + filename + '.tsv','w',newline='') as f:
        write = csv.writer(f, delimiter = '\t')

        write.writerow(fields)
        write.writerows(triples)

    print('TSV created!')

def connect_to_sql():
    mydb = mysql.connector.connect(
        host="round4",
        user="sshoc",
        password="IqazZuKaXMmSEqUl"
    )

    return mydb

def get_property(uri):
    remove_uri = uri.replace('https://rdf.ng-london.org.uk/raphael/resource/','')
    final_property = remove_uri.replace('_',' ')
    if '.' in final_property:
        final_property = str(final_property.split('.')[1])
    if 'RRR' in final_property:
        final_property = final_property.replace('RRR','')
    return final_property

def get_json(url):
    operUrl = urllib.request.urlopen(url)
    if(operUrl.getcode()==200):
       data = operUrl.read()
       json_data = json.loads(data)
    else:
       print("Error receiving data", operUrl.getcode())
    return json_data

def check_pids_csv(input_literal):
    pid_value = 'None'
    if os.path.isfile('outputs/placeholder_pids.csv') == True:
        with open('outputs/placeholder_pids.csv','r') as f:
            reader = csv.DictReader(f, delimiter=',')
            dict_list = []
            for row in reader:
                dict_list.append(row)
            for value in range(len(dict_list)):
                csv_literal = str(dict_list[value]['Literal value'])
                if csv_literal == str(input_literal):
                    pid_value = str(dict_list[value]['Placeholder PID'])
                    break
    return pid_value

def check_db(input_literal, table_name):
    cursor = db.cursor()
    pid_value = None
    input_literal = str(input_literal)
    if table_name == 'temp_pids':
        val = 'pid_value'
        str_input = 'literal_value'
    elif table_name == 'wikidata':
        val = 'wd_value'
        str_input = 'string_literal'

    query = "SELECT " + val + " FROM sshoc." + table_name + " WHERE " + str_input + " = '" + input_literal + "'"
    cursor.execute(query)
    result = cursor.fetchall()
    
    if len(result) > 0:
        pid_value = result[0][0]

    return pid_value

def check_elasticsearch(search_term, index_name): 
    es = Elasticsearch([{'host':'localhost','port':9200}]) 
    search_term = str(search_term)

    pid_value = None
    body_query = {'query':{'match':{'column1':{'query':search_term, 'fuzziness':0}}}}
    res = es.search(index=index_name,body=body_query)

    for hit in res['hits']['hits']:
        pid_value = hit['_source']['column2']
    
    return pid_value

def find_old_pid(ng_number):
    #json_data = get_json('https://scientific.ng-london.org.uk/export/ngpidexport_00A.json')
    if ng_number.startswith('https'):
        ng_number = str(ng_number.rsplit('/',1)[-1])
    else:
        ng_number = ng_number.replace(' ','_')
    old_pid = None

    try:
        export_url = 'https://collectiondata.ng-london.org.uk/es/ng-public/_search?q=identifier.object_number:' + ng_number
        json = get_json(export_url)
        if json['hits']['total'] > 0:
            old_pid = json['hits']['hits'][0]['_id']
    except:
        old_pid = None

    return old_pid

def generate_placeholder_PID(input_literal):
    N = 4
    placeholder_PID = ""
    for number in range(N):
        res = ''.join(random.choices(string.ascii_uppercase + string.digits, k = N))
        placeholder_PID += str(res)
        placeholder_PID += '-'
    placeholder_PID = placeholder_PID[:-1]
    #input_list = [input_literal, placeholder_PID]
    #fields = ['Literal value','Placeholder PID']
    existing_pid = check_db(input_literal, table_name = 'temp_pids')
    old_pid = find_old_pid(input_literal)
    
    if existing_pid is not None:
        return existing_pid
    elif old_pid is not None:
        return old_pid
    else:
        #db = connect_to_sql()
        cursor = db.cursor()
        input_literal = str(input_literal)
        query = 'INSERT INTO sshoc.temp_pids (literal_value, pid_value) VALUES (%s, %s)'
        val = (input_literal, placeholder_PID)
        cursor.execute(query, val)
        db.commit()

        placeholder_PID = str(placeholder_PID)
        return placeholder_PID

def create_PID_from_triple(pid_type, subj):
    if pid_type == 'object':
        pid_name = subj
    else:
        pid_name = pid_type + ' of ' + subj
    output_pid = generate_placeholder_PID(pid_name)

    return output_pid

def find_aat_value(material,material_type):
    material = str(get_property(material))
    if material_type == getattr(RRO,'RP20.has_medium'):
        wb = load_workbook(filename = 'inputs/NG_Medium_and_Support_AAT.xlsx', read_only=True)
        ws = wb['Medium Material']
    elif material_type == getattr(RRO,'RP32.has_support'):
        wb = load_workbook(filename = 'inputs/NG_Medium_and_Support_AAT.xlsx', read_only=True)
        ws = wb['Support Materials']
    elif material_type == 'medium type':
        wb = load_workbook(filename = 'inputs/NG_Medium_and_Support_AAT.xlsx', read_only=True)
        ws = wb['Medium Type']
    elif material_type == 'support type':
        wb = load_workbook(filename = 'inputs/NG_Medium_and_Support_AAT.xlsx', read_only=True)
        ws = wb['Support Type']
    elif material_type == getattr(RRO, 'RP215.has_acted_in_the_role_of_an'):
        wb = load_workbook(filename = 'inputs/NG_Roles_AAT.xlsx', read_only=True)
        ws = wb['AAT_Roles']
    for row in ws.iter_rows(values_only=True):
        if material in row:
            aat_number_string = str(row[1])
            aat_number = aat_number_string.replace('aat:','')
            aat_type = row[2]
            return aat_number, aat_type
    wb.close()

def wikidata_query(literal, literal_type):
    sparql = SPARQLWrapper('https://query.wikidata.org/sparql')
    print('Trying a query with input ' + literal)
    remove_uri = literal.replace('https://rdf.ng-london.org.uk/raphael/resource/','')
    literal = remove_uri.replace('_',' ')
    literal = literal.replace('%C3%A9', 'e')
    literal = literal.replace('%C3%A0', 'a')
    
    if literal_type == 'year':
        thing_type = 'Q577'
        if 'RRR' in literal:
            string_literal = str(literal.rsplit(' ')[-1])
        elif 'About' in literal:
            string_literal = str(literal.rsplit('-')[-1])
        else:
            return None
    elif literal_type == 'institution':
        thing_type = 'Q207694'
        string_literal = str(literal)
    elif literal_type == 'location':
        thing_type = 'Q515'
        string_literal = str(literal)
    elif literal_type == 'language':
        thing_type = 'Q34770'
        string_literal = str(literal)
    
    if string_literal is not None:
        result = check_db(string_literal, 'wikidata')
    else:
        result is None
    
    if result is not None:
        result = result.replace('http://www.wikidata.org/entity/','')
        return result
    else:
        query = ('''
        SELECT DISTINCT ?year
        WHERE
        {
        ?year  wdt:P31 wd:''' + thing_type + ''' .
                ?year rdfs:label ?yearLabel .
        FILTER( str(?yearLabel) = "''' + string_literal + '''" ) .
        SERVICE wikibase:label { bd:serviceParam wikibase:language "[AUTO_LANGUAGE],en". }
        }
        ''')
        
        sparql.setQuery(query)
        sparql.setReturnFormat(JSON)

        ret = sparql.query().convert()
        time.sleep(3)
        
        if ret["results"]["bindings"] != []:
            for result in ret["results"]["bindings"]:
                #db = connect_to_sql()
                cursor = db.cursor()
                result = result["year"]["value"]
                query = 'INSERT INTO sshoc.wikidata (string_literal, wd_value) VALUES (%s, %s)'
                val = (string_literal, result)
                cursor.execute(query, val)
                db.commit()
        else:
            result is None
        
        if result is not None:
            result = result.replace('http://www.wikidata.org/entity/','')

        return result

def run_ruby_program(input_string):
    import subprocess

    ruby_var = 'ruby citation_parser.rb \'' + input_string + '\''
    output = subprocess.Popen(ruby_var, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    out, error = output.communicate()

    try:
        string_output = out.decode("utf-8")
        json_output = json.loads(string_output)
    except:
        return

    return json_output

def create_year_dates(year):
    import datetime
    year = int(year)

    start = datetime.datetime(year, 1, 1)
    end = datetime.datetime(year, 12, 31)

    return start, end

def create_triples_from_reference_string(references_list, painting_id):
    for reference in references_list:
        #reference_json = run_ruby_program(reference)
        reference_PID = BNode()
        #parse_reference_json(reference_json, reference_PID)
        painting_PID = generate_placeholder_PID(painting_id)

        new_graph.add((getattr(NGO, reference_PID), CRM.P67_refers_to, getattr(NGO, painting_PID)))
        new_graph.add((getattr(NGO, reference_PID), RDF.type, CRM.E31_Document))
        new_graph.add((getattr(NGO, reference_PID), CRM.P2_has_type, CRM.E31_Document))
        new_graph.add((getattr(NGO, reference_PID), RDFS.label, Literal(reference)))

    return new_graph

def create_title_triples(PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP34.has_title'):
        obj = obj.replace("'","")
        title_PID = generate_placeholder_PID(obj)

        new_graph.add((getattr(NGO,PID), CRM.P102_has_title, getattr(NGO, title_PID)))
        new_graph.add((getattr(NGO, title_PID), RDFS.label, Literal(obj, lang="en")))
        new_graph.add((getattr(NGO, title_PID), RDF.type, CRM.E35_Title))
        new_graph.add((getattr(NGO, title_PID), CRM.P2_has_type, CRM.E35_Title))
        new_graph.add((getattr(NGO, title_PID), CRM.P2_has_type, getattr(AAT,'300417209')))
        new_graph.add((getattr(AAT,'300417209'), RDFS.label, Literal('full title', lang="en")))

    elif pred == getattr(RRO, 'RP31.has_short_title'):
        obj = obj.replace("'","")
        title_PID = generate_placeholder_PID(obj)

        new_graph.add((getattr(NGO,PID), CRM.P102_has_title, getattr(NGO, title_PID)))
        new_graph.add((getattr(NGO, title_PID), RDFS.label, Literal(obj, lang="en")))
        new_graph.add((getattr(NGO, title_PID), RDF.type, CRM.E35_Title))
        new_graph.add((getattr(NGO, title_PID), CRM.P2_has_type, CRM.E35_Title))
        new_graph.add((getattr(NGO, title_PID), CRM.P2_has_type, getattr(AAT,'300417208')))
        new_graph.add((getattr(AAT,'300417208'), RDFS.label, Literal('brief title', lang="en")))

    return new_graph

def create_medium_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP20.has_medium'):
        medium_PID = create_PID_from_triple('medium', subj)
        medium_BN = BNode()
        m2_bn = BNode()
        aat_number, aat_type = find_aat_value(obj, pred)
        aat_classification_number, aat_classification_type = find_aat_value(obj, 'medium type')

        new_graph.add((getattr(NGO,subject_PID), CRM.P46_is_composed_of, getattr(NGO, medium_PID)))
        new_graph.add((getattr(NGO, medium_PID), CRM.P45_consists_of, medium_BN))
        new_graph.add((medium_BN, CRM.P2_has_type, getattr(AAT,aat_number)))
        new_graph.add(((getattr(AAT,aat_number)), RDFS.label, Literal(aat_type)))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT,aat_classification_number)))
        new_graph.add(((getattr(AAT,aat_classification_number)), RDFS.label, Literal(aat_classification_type)))
        new_graph.add((getattr(NGO, medium_PID), RDF.type, CRM.E24_Physical_Human_Made_Thing))
        new_graph.add((getattr(NGO, medium_PID), CRM.P2_has_type, CRM.E24_Physical_Human_Made_Thing))
        new_graph.add((getattr(NGO, medium_PID), CRM.P2_has_type, getattr(AAT,'300163343')))
        new_graph.add(((getattr(AAT,'300163343')), RDFS.label, Literal('media (artists\' material)', lang="en")))

    elif pred == getattr(RRO, 'RP32.has_support'):
        medium_PID = create_PID_from_triple('support', subj)
        medium_BN = BNode()
        aat_number, aat_type = find_aat_value(obj, pred)
        aat_classification_number, aat_classification_type = find_aat_value(obj, 'support type')
        new_graph.add((getattr(NGO,subject_PID), CRM.P46_is_composed_of, getattr(NGO, medium_PID)))
        new_graph.add((getattr(NGO, medium_PID), CRM.P45_consists_of, medium_BN))
        new_graph.add((medium_BN, CRM.P2_has_type, getattr(AAT,aat_number)))
        new_graph.add(((getattr(AAT,aat_number)), RDFS.label, Literal(aat_type)))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT,aat_classification_number)))
        new_graph.add(((getattr(AAT,aat_classification_number)), RDFS.label, Literal(aat_classification_type)))
        new_graph.add((getattr(NGO, medium_PID), RDF.type, CRM.E24_Physical_Human_Made_Thing))
        new_graph.add((getattr(NGO, medium_PID), CRM.P2_has_type, CRM.E24_Physical_Human_Made_Thing))
        new_graph.add((getattr(NGO, medium_PID), CRM.P2_has_type, getattr(AAT,'300014844')))
        new_graph.add(((getattr(AAT,'300014844')), RDFS.label, Literal('supports (artists\' materials)', lang="en")))

    return new_graph

def create_collection_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP99.is_part_of'):
        collection_PID = create_PID_from_triple('object', obj)
        literal_obj = obj.replace('https://rdf.ng-london.org.uk/raphael/resource/','').replace('_',' ')

        new_graph.add((getattr(NGO,subject_PID), CRM.P46i_forms_part_of, getattr(NGO, collection_PID)))
        new_graph.add((getattr(NGO, collection_PID), RDF.type, CRM.E78_Curated_Holding))
        new_graph.add((getattr(NGO, collection_PID), CRM.P2_has_type, CRM.E78_Curated_Holding))
        new_graph.add((getattr(NGO, collection_PID), RDFS.label, Literal(literal_obj, lang="en")))
        new_graph.add((getattr(NGO, collection_PID), CRM.P2_has_type, getattr(AAT, '300443858')))
        new_graph.add((getattr(AAT, '300443858'), RDFS.label, Literal('collections (repositories)', lang="en")))

    return new_graph

def create_dimension_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP36.has_width_in_cm'):
        dimension_PID = BNode()

        new_graph.add((getattr(NGO,subject_PID), CRM.P43_has_dimension, dimension_PID))
        new_graph.add((dimension_PID, RDF.type, CRM.E54_Dimension))
        new_graph.add((dimension_PID, CRM.P2_has_type, CRM.E54_Dimension))
        if ';' in obj:
            new_graph.add((dimension_PID, CRM.P90_has_value, Literal(obj, datatype=XSD.string)))
        else:
            new_graph.add((dimension_PID, CRM.P90_has_value, Literal(obj, datatype=XSD.double)))
        new_graph.add((dimension_PID, CRM.P91_has_unit, getattr(AAT,'300379098')))
        new_graph.add((getattr(AAT,'300379098'), RDFS.label, Literal('centimeters', lang="en")))
        new_graph.add((dimension_PID, CRM.P2_has_type, getattr(AAT,'300055647')))
        new_graph.add((getattr(AAT,'300055647'), RDFS.label, Literal('width', lang="en")))

    elif pred == getattr(RRO, 'RP16.has_height_in_cm'): 
        dimension_PID = BNode()

        new_graph.add((getattr(NGO,subject_PID), CRM.P43_has_dimension, dimension_PID))
        new_graph.add((dimension_PID, RDF.type, CRM.E54_Dimension))
        new_graph.add((dimension_PID, CRM.P2_has_type, CRM.E54_Dimension))
        if ';' in obj:
            new_graph.add((dimension_PID, CRM.P90_has_value, Literal(obj, datatype=XSD.string)))
        else:
            new_graph.add((dimension_PID, CRM.P90_has_value, Literal(obj, datatype=XSD.double)))
        new_graph.add((dimension_PID, CRM.P91_has_unit, getattr(AAT,'300379098')))
        new_graph.add((getattr(AAT,'300379098'), RDFS.label, Literal('centimeters', lang="en")))
        new_graph.add((dimension_PID, CRM.P2_has_type, getattr(AAT,'300055644')))
        new_graph.add((getattr(AAT,'300055644'), RDFS.label, Literal('height', lang="en")))

    elif pred == getattr(RRO, 'RP225.has_width_in_pixels'):
        dimension_PID = BNode()

        new_graph.add((getattr(NGO, subject_PID), DIG.L56_has_pixel_width, Literal(obj, datatype=XSD.double)))

    elif pred == getattr(RRO, 'RP227.has_height_in_pixels'):
        dimension_PID = BNode()

        new_graph.add((getattr(NGO, subject_PID), DIG.L57_has_pixel_height, Literal(obj, datatype=XSD.double)))

    return new_graph

def create_identifier_triples(subject_PID, pred, obj):
    if pred == getattr(RRO, 'RP17.has_identifier'):
        blank_node = BNode()
        new_graph.add((getattr(NGO, subject_PID), CRM.P48_has_preferred_identifier, blank_node))
        new_graph.add((blank_node, RDF.type, CRM.E42_Identifier))
        new_graph.add((blank_node, CRM.P2_has_type, CRM.E42_Identifier))
        new_graph.add((blank_node, CRM.P2_has_type, getattr(AAT, '300404626')))
        new_graph.add((getattr(AAT, '300404626'), RDFS.label, Literal('identification numbers', lang="en")))
        new_graph.add((blank_node, RDFS.label, Literal(obj, lang="en")))

    return new_graph

def create_type_triples(subject_PID, pred, obj):
    if pred == RDF.type and obj == getattr(RRO, 'RC12.Painting'):
        new_graph.add((getattr(NGO, subject_PID), RDF.type, CRM.E24_Physical_Human_Made_Thing))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, CRM.E24_Physical_Human_Made_Thing))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT, '300033618')))
        new_graph.add((getattr(AAT, '300033618'), RDFS.label, Literal('paintings (visual works)', lang="en")))
    elif pred == RDF.type and obj == getattr(RRO, 'RC40.Person'):
        new_graph.add((getattr(NGO, subject_PID), RDF.type, CRM.E21_Person))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, CRM.E21_Person))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT, '300024979')))
        new_graph.add((getattr(AAT, '300024979'), RDFS.label, Literal('people (agents)', lang="en")))
    elif pred == RDF.type and obj == getattr(RRO, 'RC41.Institution'):
        new_graph.add((getattr(NGO, subject_PID), RDF.type, CRM.E74_Group))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, CRM.E74_Group))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT, '300026004')))
        new_graph.add((getattr(AAT, '300026004'), RDFS.label, Literal('institutions (organizations)', lang="en")))
    elif pred == RDF.type and obj == getattr(RRO, 'RC10.Building'):
        new_graph.add((getattr(NGO, subject_PID), RDF.type, CRM.E53_Place))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, CRM.E53_Place))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT, '300312292')))
        new_graph.add((getattr(AAT, '300312292'), RDFS.label, Literal('institutions (buildings)', lang="en")))
    elif pred == RDF.type and (obj in [getattr(RRO, 'RC26.Digital_Document'), getattr(RRO, 'RC220.Digital_Text')]):
        new_graph.add((getattr(NGO, subject_PID), RDF.type, DIG.D1_Digital_Object))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, DIG.D1_Digital_Object))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT, '300424602')))
        new_graph.add((getattr(AAT, '300424602'), RDFS.label, Literal('digital document', lang="en")))
    elif pred == RDF.type and obj == getattr(RRO, 'RC25.Image'):
        new_graph.add((getattr(NGO, subject_PID), RDF.type, DIG.D1_Digital_Object))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, DIG.D1_Digital_Object))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT, '300215302')))
        new_graph.add((getattr(AAT, '300215302'), RDFS.label, Literal('digital images', lang="en")))

    return new_graph

def create_time_span_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP209.has_time-span'):
        time_span_PID = BNode()

        #will need to change this logic if we encounter a time span that isn't a year
        new_graph.add((getattr(NGO, subject_PID), CRM.P4_has_time_span, time_span_PID))
        new_graph.add((time_span_PID, RDF.type, CRM.E52_Time_span))
        new_graph.add((time_span_PID, CRM.P2_has_type, CRM.E52_Time_span))
        new_graph.add((time_span_PID, CRM.P2_has_type, getattr(AAT, '300379244')))
        new_graph.add((getattr(AAT, '300379244'), RDFS.label, Literal('years', lang="en")))
        new_graph.add((time_span_PID, RDFS.label, Literal(obj, lang="en")))

        obj_year = get_property(obj)
        api_link = 'https://scientific.ng-london.org.uk/api/api-tms-v.2.3.0.php?format=json&what=timespan&which=' + obj_year
        r = requests.get(api_link)
        response_json = r.json()
        start_date = response_json[1][0]
        end_date = response_json[1][1]
        
        wikidata_year = wikidata_query(obj_year, 'year')
        new_graph.add((time_span_PID, CRM.P82a_begin_of_the_begin, Literal(start_date, datatype=XSD.dateTime)))
        new_graph.add((time_span_PID, CRM.P82b_end_of_the_end, Literal(end_date, datatype=XSD.dateTime)))
        if wikidata_year != None and wikidata_year != 'No WD value':
            new_graph.add((time_span_PID, OWL.sameAs, getattr(WD, wikidata_year)))
        
    return new_graph

def create_event_triples(subject_PID, obj_PID, subj, pred):
    if pred == getattr(RRO, 'RP68.was_acquired'):
        event_type = CRM.E8_Acquisition
        event_property = CRM.P24i_changed_ownership_through
        aat_event_id = getattr(AAT, '300157782')
        aat_event_type = Literal('acquisition (collections management)', lang="en")
    elif pred == getattr(RRO, 'RP72.was_produced'):
        event_type = CRM.E12_Production
        event_property = CRM.P108i_was_produced_by
        aat_event_id = getattr(AAT, '300404387')
        aat_event_type = Literal('creating (artistic activity)', lang="en")
    elif pred == getattr(RRO, 'RP42.was_born_in'):
        event_type = CRM.E67_Birth
        event_property = CRM.P98i_was_born
        aat_event_id = getattr(AAT, '300069672')
        aat_event_type = Literal('births', lang="en")
    elif pred == getattr(RRO, 'RP4.died_in'):
        event_type = CRM.E69_Death
        event_property = CRM.P100i_died_in
        aat_event_id = getattr(AAT, '300151836')
        aat_event_type = Literal('deaths', lang="en")

    related_painting_history_event = 'History of ' + subj
    related_painting_history_event_PID = create_PID_from_triple('history', subj)

    new_graph.add((getattr(NGO, subject_PID), event_property, getattr(NGO, obj_PID)))
    new_graph.add((getattr(NGO, obj_PID), RDF.type, event_type))
    new_graph.add((getattr(NGO, obj_PID), CRM.P2_has_type, event_type))
    new_graph.add((getattr(NGO, obj_PID), CRM.P2_has_type, aat_event_id))
    new_graph.add((aat_event_id, RDFS.label, aat_event_type))
    new_graph.add((getattr(NGO, obj_PID), CRM.P10_falls_within, getattr(NGO, related_painting_history_event_PID)))
    new_graph.add((getattr(RRO, related_painting_history_event_PID), RDF.type, CRM.E5_Event))
    new_graph.add((getattr(RRO, related_painting_history_event_PID), CRM.P2_has_type, CRM.E5_Event))
    new_graph.add((getattr(RRO, related_painting_history_event_PID), CRM.P2_has_type, getattr(AAT, '300055863')))
    new_graph.add((getattr(AAT, '300055863'), RDFS.label, Literal('provenance (history of ownership)', lang="en")))
    new_graph.add((getattr(RRO, related_painting_history_event_PID), RDFS.label, Literal(related_painting_history_event, lang="en")))

    return new_graph

def create_name_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP56.has_name'):
        name_PID = BNode()
        new_graph.add((getattr(NGO, subject_PID), CRM.P1_is_identified_by, name_PID))
        new_graph.add((name_PID, RDF.type, CRM.E41_Appellation))
        new_graph.add((name_PID, CRM.P2_has_type, CRM.E41_Appellation))
        new_graph.add((name_PID, RDFS.label, Literal(obj, lang="en")))
        new_graph.add((getattr(NGO, name_PID), CRM.P2_has_type, getattr(AAT, '300404688')))
        new_graph.add((getattr(AAT, '300404688'), RDFS.label, Literal('full names', lang="en")))
    elif pred == getattr(RRO, 'RP11.has_current_position'):
        job_title_PID = BNode()
        new_graph.add((getattr(NGO, subject_PID), CRM.P1_is_identified_by, job_title_PID))
        new_graph.add((job_title_PID, RDF.type, CRM.E41_Appellation))
        new_graph.add((job_title_PID, CRM.P2_has_type, CRM.E41_Appellation))
        new_graph.add((job_title_PID, RDFS.label, Literal(obj, lang="en")))
        new_graph.add((getattr(NGO, job_title_PID), CRM.P2_has_type, getattr(AAT, '300263369')))
        new_graph.add((getattr(AAT, '300263369'), RDFS.label, Literal('occupations (livelihoods)', lang="en")))

    return new_graph

def create_role_triples(subject_PID, pred, obj):
    if pred == getattr(RRO, 'RP215.has_acted_in_the_role_of_an'):
        aat_number, aat_type = find_aat_value(obj, pred)

        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT, aat_number)))
        new_graph.add((getattr(AAT, aat_number), RDFS.label, Literal(aat_type)))

    return new_graph

def create_comment_triples(subject_PID, pred, obj):
    if pred == getattr(RRO, 'RP59.has_description') or pred == getattr(RRO, 'RP237.has_content'):
        new_graph.add((getattr(NGO, subject_PID), CRM.P3_has_note, Literal(obj)))

    return new_graph

def create_location_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP10.has_current_location') or pred == getattr(RRO, 'RP90.is_located_within'):
        try:
            location_PID = wikidata_query(obj, 'location')
        except:
            location_PID = generate_placeholder_PID(obj)
        if location_PID is None:
            location_PID = generate_placeholder_PID(obj)
        literal_obj = obj.replace('https://rdf.ng-london.org.uk/raphael/resource/','').replace('_',' ')
        
        if location_PID.startswith('http://www.wikidata.org'):
            new_graph.add((getattr(NGO, subject_PID), CRM.P53_has_former_or_current_location, getattr(WD, location_PID)))
            new_graph.add((getattr(WD, location_PID), RDF.type, CRM.E53_Place))
            new_graph.add((getattr(WD, location_PID), CRM.P2_has_type, CRM.E53_Place))
            new_graph.add((getattr(WD, location_PID), RDFS.label, Literal(literal_obj, lang="en")))
        else:
            new_graph.add((getattr(NGO, subject_PID), CRM.P53_has_former_or_current_location, getattr(NGO, location_PID)))
            new_graph.add((getattr(NGO, location_PID), RDF.type, CRM.E53_Place))
            new_graph.add((getattr(NGO, location_PID), CRM.P2_has_type, CRM.E53_Place))
            new_graph.add((getattr(NGO, location_PID), RDFS.label, Literal(literal_obj, lang="en")))

    return new_graph

def create_actor_event_relationship_triples(subject_PID, pred, obj):
    if pred == getattr(RRO, 'RP43.was_carried_out_by'):
        actor_PID = create_PID_from_triple('object', obj)

        new_graph.add((getattr(NGO, subject_PID), CRM.P14_carried_out_by, getattr(NGO, actor_PID)))
        new_graph.add((getattr(NGO, actor_PID), RDF.type, CRM.E39_Actor))
        new_graph.add((getattr(NGO, actor_PID), CRM.P2_has_type, CRM.E39_Actor))
        new_graph.add((getattr(NGO, actor_PID), RDFS.label, Literal(obj, lang="en")))
        new_graph.add((getattr(NGO, actor_PID), CRM.P2_has_type, getattr(AAT, '300024979')))
        new_graph.add((getattr(AAT, '300024979'), RDFS.label, Literal('people (agents)', lang="en")))
    if pred == getattr(RRO, 'RP9.has_curator'):
        curator_PID = generate_placeholder_PID(obj)

        new_graph.add((getattr(NGO, subject_PID), CRM.P109_has_current_or_former_curator, getattr(NGO, curator_PID)))
        new_graph.add((getattr(NGO, curator_PID), RDF.type, CRM.E39_Actor))
        new_graph.add((getattr(NGO, curator_PID), CRM.P2_has_type, CRM.E39_Actor))
        new_graph.add((getattr(NGO, curator_PID), RDFS.label, Literal(obj, lang="en")))
        new_graph.add((getattr(NGO, curator_PID), CRM.P2_has_type, getattr(AAT, '300024979')))
        new_graph.add((getattr(AAT, '300024979'), RDFS.label, Literal('people (agents)', lang="en")))
    if pred == getattr(RRO, 'RP201.is_current_keeper_of'):
        collection_PID = create_PID_from_triple('object', obj)
        literal_obj = obj.replace('https://rdf.ng-london.org.uk/raphael/resource/','').replace('_',' ')

        new_graph.add((getattr(NGO, subject_PID), CRM.P50i_is_current_keeper_of, getattr(NGO, collection_PID)))
        new_graph.add((getattr(NGO,collection_PID), RDF.type, CRM.E78_Curated_Holding))
        new_graph.add((getattr(NGO,collection_PID), CRM.P2_has_type, CRM.E78_Curated_Holding))
        new_graph.add((getattr(NGO,collection_PID), RDFS.label, Literal(literal_obj, lang="en")))
        new_graph.add((getattr(NGO, collection_PID), CRM.P2_has_type, getattr(AAT, '300443858')))
        new_graph.add((getattr(AAT, '300443858'), RDFS.label, Literal('collections (repositories)', lang="en")))
    if pred == getattr(RRO, 'RP40.is_related_to'):
        object_PID = generate_placeholder_PID(obj)

        new_graph.add((getattr(NGO, subject_PID), CRM.P73_refers_to, getattr(NGO, object_PID)))

    return new_graph

def create_documentation_triples(subject_PID, pred, obj):
    if pred == getattr(RRO, 'RP245.has_website') or pred == getattr(RRO, 'RP257.has_external_link'):
        website_PID = BNode()

        new_graph.add((getattr(NGO, subject_PID), CRM.P70i_is_documented_in, website_PID))
        new_graph.add((website_PID, RDF.type, CRM.E73_Information_Object))
        new_graph.add((website_PID, CRM.P2_has_type, CRM.E73_Information_Object))
        new_graph.add((website_PID, RDFS.label, Literal(obj, lang="en")))
        new_graph.add((website_PID, CRM.P2_has_type, getattr(AAT, '300265431')))
        new_graph.add((getattr(AAT, '300265431'), RDFS.label, Literal('Web sites', lang="en")))

    return new_graph

def create_institution_triples(old_graph, new_graph, subject_PID, institution_name):
    for subj, pred, obj in old_graph.triples((institution_name, None, None)):
        
        if obj != getattr(RRO, 'RC10.Building'):
            new_graph = create_type_triples(subject_PID, pred, obj)
            new_graph = create_comment_triples(subject_PID, pred, obj)
            new_graph = create_actor_event_relationship_triples(subject_PID, pred, obj)
            new_graph = create_role_triples(subject_PID, pred, obj)
            new_graph = create_documentation_triples(subject_PID, pred, obj)
            new_graph = create_collection_triples(subject_PID, subj, pred, obj)
            new_graph = create_identifier_triples(subject_PID, pred, obj)

        elif obj == getattr(RRO, 'RC10.Building'):
            object_PID = create_PID_from_triple('building', subj)
            new_graph.add((getattr(NGO, subject_PID), CRM.P156_occupies, getattr(NGO, object_PID)))
            new_graph = create_building_triples(old_graph, new_graph, object_PID, subj)

    return new_graph

def create_building_triples(old_graph, new_graph, subject_PID, building_name):
    for subj, pred, obj in old_graph.triples((building_name, None, None)):
        if obj != getattr(RRO, 'RC41.Institution'):
            new_graph = create_location_triples(subject_PID, subj, pred, obj)
            new_graph = create_type_triples(subject_PID, pred, obj)
            new_graph = create_identifier_triples(subject_PID, pred, obj)

    return new_graph

def create_room_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP90.is_located_within'):
        obj_PID = create_PID_from_triple('building', obj)
        literal_subj = subj.replace('https://rdf.ng-london.org.uk/raphael/resource/','').replace('_',' ')

        new_graph.add((getattr(NGO, subject_PID), RDFS.label, Literal(literal_subj, lang="en")))
        new_graph.add((getattr(NGO, subject_PID), CRM.P89_falls_within, getattr(NGO, obj_PID)))
        new_graph.add((getattr(NGO, subject_PID), RDF.type, CRM.E53_Place))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, CRM.E53_Place))

    return new_graph

def create_area_of_room_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP90_is_located_within'):
        obj_PID = generate_placeholder_PID(obj)
        new_graph.add((getattr(NGO, subject_PID), RDFS.label, Literal(obj, lang="en")))
        new_graph.add((getattr(NGO, subject_PID), CRM.P89_falls_within, getattr(NGO, obj_PID)))
        new_graph.add((getattr(NGO, subject_PID), RDF.type, CRM.E53_Place))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, CRM.E53_Place))

    return new_graph

def parse_reference_json(reference_json, subject_PID):
    creation_event = BNode()
    title = reference_json[0]["title"][0]
    title_PID = BNode()
    date = reference_json[0]["date"][0]
    journal_title = reference_json[0]["container-title"][0]
    journal_title_PID = BNode()
    start_date, end_date = create_year_dates(date)
    time_span_PID = BNode()

    for i in range(0, len(reference_json[0]["author"])):
        author_node = BNode()
        author = reference_json[0]["author"][i]["given"] + " " + reference_json[0]["author"][i]["family"]
        new_graph.add((creation_event, CRM.P14_carried_out_by, author_node))
        new_graph.add((author_node, RDF.type, CRM.E39_Actor))
        new_graph.add((author_node, CRM.P2_has_type, CRM.E39_Actor))
        new_graph.add((author_node, RDFS.label, Literal(author, lang="en")))

    new_graph.add((getattr(NGO, subject_PID), CRM.P102_has_title, title_PID))
    new_graph.add((title_PID, RDFS.label, Literal(title, lang="en")))
    new_graph.add((title_PID, CRM.P106i_forms_part_of, journal_title_PID))
    new_graph.add((journal_title_PID, RDFS.label, Literal(journal_title, lang="en")))
    new_graph.add((getattr(NGO, subject_PID), CRM.P94i_was_created_by, creation_event))
    new_graph.add((creation_event, RDF.type, CRM.E65_Creation))
    new_graph.add((creation_event, CRM.P2_has_type, CRM.E65_Creation))

    new_graph.add((creation_event, CRM.P4_has_time_span, time_span_PID))
    new_graph.add((time_span_PID, RDF.type, CRM.E52_Time_span))
    new_graph.add((time_span_PID, CRM.P2_has_type, CRM.E52_Time_span))
    new_graph.add((time_span_PID, CRM.P2_has_type, getattr(AAT, '300379244')))
    new_graph.add((getattr(AAT, '300379244'), RDFS.label, Literal('years', lang="en")))
    wikidata_year = wikidata_query(date, 'year')
    new_graph.add((time_span_PID, CRM.P82a_begin_of_the_begin, Literal(start_date, datatype=XSD.dateTime)))
    new_graph.add((time_span_PID, CRM.P82b_end_of_the_end, Literal(end_date, datatype=XSD.dateTime)))
    if wikidata_year != None and wikidata_year != 'No WD value':
        new_graph.add((time_span_PID, OWL.sameAs, getattr(WD, wikidata_year)))

def create_reference_triples(subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP84.has_reference') or pred == getattr(RRO, 'RP233.has_caption'):
        if str(obj).startswith('Conservation Dossier') == True:
            ref = str(obj)
            ref_array = ref.split()
            
            painting_id = ref_array[2].replace(',','')
            dossier_volume = ref_array[3].replace(',','')
            file_PID = subject_PID

            related_painting_PID = generate_placeholder_PID(painting_id)
            related_painting_history_event_PID = create_PID_from_triple('provenance', painting_id)
            related_painting_history_event = 'Provenance of ' + painting_id
            dossier_title = 'Conservation Dossier of ' + painting_id
            dossier_PID = generate_placeholder_PID(dossier_title)
            volume_title = 'Volume ' + dossier_volume + ' of ' + dossier_title
            volume_PID = generate_placeholder_PID(volume_title)
            page_number_value = BNode()
            volume_number_value = BNode()

            new_graph.add((getattr(NGO, dossier_PID), RDF.type, CRM.E31_Document))
            new_graph.add((getattr(NGO, dossier_PID), CRM.P2_has_type, CRM.E31_Document))
            new_graph.add((getattr(NGO, dossier_PID), RDFS.label, Literal(dossier_title, lang="en")))
            new_graph.add((getattr(NGO, dossier_PID), CRM.P148_has_component, getattr(NGO, volume_PID)))
            new_graph.add((getattr(NGO, volume_PID), RDF.type, CRM.E31_Document))
            new_graph.add((getattr(NGO, volume_PID), CRM.P2_has_type, CRM.E31_Document))
            new_graph.add((getattr(NGO, volume_PID), RDFS.label, Literal(volume_title, lang="en")))
            new_graph.add((getattr(NGO, volume_PID), CRM.P90_has_value, volume_number_value))
            new_graph.add((volume_number_value, RDF.value, Literal(dossier_volume)))
            new_graph.add((volume_number_value, CRM.P2_has_type, getattr(WD, 'P478')))
            new_graph.add((getattr(WD, 'P478'), RDFS.label, Literal('volume', lang="en")))       
            new_graph.add((page_number_value, CRM.P2_has_type, getattr(WD, 'Q11325816')))
            new_graph.add((getattr(WD, 'Q11325816'), RDFS.label, Literal('page number', lang="en")))
            new_graph.add((getattr(NGO, dossier_PID), CRM.P70_documents, getattr(NGO, related_painting_history_event_PID)))
            new_graph.add((getattr(NGO, related_painting_history_event_PID), RDFS.label, Literal(related_painting_history_event, lang="en")))

            if len(ref_array) >= 7:
                vol_page_no = ref_array[5]
                volume_page_name = 'Page ' + vol_page_no + ' of ' + volume_title
                page_PID = generate_placeholder_PID(volume_page_name)
                dossier_page_no = ref_array[7].replace(').','')
                dossier_page_name = 'Page ' + dossier_page_no + ' of ' + dossier_title
                new_graph.add((getattr(NGO, page_PID), RDFS.label, Literal(dossier_page_name, lang="en")))
                new_graph.add((getattr(NGO, page_PID), CRM.P67_refers_to, getattr(NGO, related_painting_PID)))

            if (len(ref_array) > 4) or (ref_array[3] != 'envelope.'):
                if (ref_array[4] != 'front.' and ref_array[4] != 'back.' and ref_array[4] != '(envelope).'):
                    vol_page_no = ref_array[5]
                    volume_page_name = 'Page ' + vol_page_no + ' of ' + volume_title
                    page_PID = generate_placeholder_PID(volume_page_name)
                    new_graph.add((getattr(NGO, page_PID), RDFS.label, Literal(volume_page_name, lang="en")))
                    new_graph.add((getattr(NGO, page_PID), CRM.P48_has_preferred_identifier, Literal(volume_page_name, lang="en")))
                    new_graph.add((getattr(NGO, page_PID), CRM.P90_has_value, page_number_value))
                    new_graph.add((getattr(NGO, page_PID), RDF.type, CRM.E31_Document))
                    new_graph.add((getattr(NGO, page_PID), CRM.P2_has_type, CRM.E31_Document))
                    new_graph.add((getattr(NGO, page_PID), CRM.P138i_has_representation, getattr(NGO, file_PID)))
                    new_graph.add((page_number_value, RDF.value, Literal(vol_page_no)))
                    new_graph.add((getattr(NGO, volume_PID), CRM.P148_has_component, getattr(NGO, page_PID)))

        else: 
            reference_json = run_ruby_program(obj)
            try:
                parse_reference_json(reference_json, subject_PID)
            except:
                return new_graph
        
    return new_graph

def create_file_triples(old_graph, subject_PID, subj, pred, obj):
    file_PID = subject_PID
    related_work = query_objects(old_graph, subj, getattr(RRO, 'RP40.is_related_to'), None)

    if pred == getattr(RRO, 'RP95.has_file_name'):
        file_name = BNode()
        new_graph.add((getattr(NGO, subject_PID), CRM.P149_is_identified_by, file_name))
        new_graph.add((file_name, RDF.type, CRM.E42_Identifier))
        new_graph.add((file_name, CRM.P2_has_type, CRM.E42_Identifier))
        new_graph.add((file_name, CRM.P2_has_type, getattr(WD, 'Q1144928')))
        new_graph.add((getattr(WD, 'Q1144928'), RDFS.label, Literal('filename', lang="en")))
        new_graph.add((file_name, RDFS.label, Literal(obj, lang="en")))
        new_graph.add((getattr(NGO, subject_PID), RDF.type, DIG.D1_Digital_Object))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, DIG.D1_Digital_Object))

    elif pred == getattr(RRO, 'RP15.has_format'):
        if obj == getattr(RRI, 'RCL88.PDF'):
            new_graph.add((getattr(NGO, file_PID), CRM.P2_has_type, getattr(AAT, '300266022')))
            new_graph.add((getattr(AAT, '300266022'), RDFS.label, Literal('PDF', lang="en")))
        elif obj == getattr(RRI, 'RCL90.Jpeg'):
            new_graph.add((getattr(NGO, file_PID), CRM.P2_has_type, getattr(AAT, '300266224')))
            new_graph.add((getattr(AAT, '300266224'), RDFS.label, Literal('JPEG', lang="en")))
        elif obj == getattr(RRI, 'RCL89.Tiff'):
            new_graph.add((getattr(NGO, file_PID), CRM.P2_has_type, getattr(AAT, '300266226')))
            new_graph.add((getattr(AAT, '300266226'), RDFS.label, Literal('TIFF', lang="en")))

    elif pred == getattr(RRO, 'RP14.has_file_size'):
        file_size = BNode()

        new_graph.add((getattr(NGO, subject_PID), CRM.P43_has_dimension, file_size))
        new_graph.add((file_size, RDF.type, CRM.E54_Dimension))
        new_graph.add((file_size, CRM.P2_has_type, CRM.E54_Dimension))
        new_graph.add((file_size, CRM.P90_has_value, Literal(obj, datatype=XSD.double)))
        new_graph.add((file_size, CRM.P91_has_unit, getattr(AAT,'300265869')))
        new_graph.add((getattr(AAT,'300265869'), RDFS.label, Literal('bytes', lang="en")))
        new_graph.add((file_size, CRM.P2_has_type, getattr(AAT,'300265863')))
        new_graph.add((getattr(AAT,'300265863'), RDFS.label, Literal('size for computer files', lang="en")))

    elif pred == getattr(RRO, 'RP30.has_pyramid'):
        if obj == 'Not Public':
            return new_graph
        else:
            pyramid = BNode()
            no_of_levels = BNode()
            no_of_levels_value = query_objects(old_graph, subj, getattr(RRO, 'RP86.has_no_of_pyramidal_levels'), None)[0]
            try:
                server_url = query_objects(old_graph, subj, getattr(RRO, 'RP243.has_pyramid_server'), None)[0]
            except:
                server_url = None
            file_path = query_objects(old_graph, subj, getattr(RRO, 'RP30.has_pyramid'), None)[0]
            if server_url is not None:
                full_file_path = server_url + "/" + file_path
            else:
                full_file_path = file_path

            file_path_bnode = BNode()

            derivation_event = create_PID_from_triple('pyramid creation for', subj)
            pyramid_PID = create_PID_from_triple('pyramid of', subj)
            pyramid_ID = BNode()
            server_PID = generate_placeholder_PID('IIIF Server')

            new_graph.add((getattr(NGO, derivation_event), RDF.type, DIG.D3_Formal_Derivation))
            new_graph.add((getattr(NGO, derivation_event), CRM.P2_has_type, DIG.D3_Formal_Derivation))
            new_graph.add((getattr(NGO, derivation_event), DIG.L21_used_as_derivation_source, getattr(NGO, subject_PID)))
            new_graph.add((getattr(NGO, derivation_event), DIG.L22_created_derivative, getattr(NGO, pyramid_PID)))
            new_graph.add((getattr(NGO, pyramid_PID), CRM.P2_has_type, pyramid))
            new_graph.add((pyramid, CRM.P2_has_type, getattr(WD, 'Q3411251')))
            new_graph.add((getattr(WD, 'Q3411251'), RDFS.label, Literal('pyramid', lang="en")))
            new_graph.add((pyramid, CRM.P43_has_dimension, no_of_levels))
            new_graph.add((no_of_levels, RDF.type, CRM.E54_Dimension))
            new_graph.add((no_of_levels, CRM.P2_has_type, CRM.E54_Dimension))
            new_graph.add((no_of_levels, CRM.P90_has_value, Literal(no_of_levels_value, datatype=XSD.double)))
            new_graph.add((no_of_levels, RDFS.label, Literal('Number of pyramidal levels')))
            new_graph.add((getattr(NGO, derivation_event), DIG.L23_used_software_or_firmware, getattr(NGO, server_PID)))
            new_graph.add((getattr(NGO, server_PID), RDF.type, DIG.D14_Software))
            new_graph.add((getattr(NGO, server_PID), CRM.P2_has_type, DIG.D14_Software))
            new_graph.add((getattr(NGO, server_PID), CRM.P2_has_type, getattr(AAT, '300266043')))
            new_graph.add((getattr(AAT, '300266043'), RDFS.label, Literal('servers (computer)', lang="en")))

            new_graph.add((getattr(NGO, pyramid_PID), CRM.P149_is_identified_by, file_path_bnode))
            new_graph.add((file_path_bnode, RDF.type, CRM.E42_Identifier))
            new_graph.add((file_path_bnode, CRM.P2_has_type, CRM.E42_Identifier))
            new_graph.add((file_path_bnode, CRM.P2_has_type, getattr(WD, 'Q1144928')))
            new_graph.add((getattr(WD, 'Q1144928'), RDFS.label, Literal('filename', lang="en")))
            new_graph.add((file_path_bnode, RDFS.label, Literal(full_file_path, lang="en")))
            new_graph.add((getattr(NGO, subject_PID), RDF.type, DIG.D1_Digital_Object))
            new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, DIG.D1_Digital_Object))

    elif pred == getattr(RRO, 'RP259.has_thumbnail'):
        thumbnail_PID = create_PID_from_triple('thumbnail of', subj)
        derivation_event = create_PID_from_triple('thumbnail creation for', subj)
        thumbnail_ID = BNode()
        server_PID = generate_placeholder_PID('IIIF Server')
        pyramid_PID = create_PID_from_triple('pyramid of', subj)

        new_graph.add((getattr(NGO, derivation_event), RDF.type, DIG.D3_Formal_Derivation))
        new_graph.add((getattr(NGO, derivation_event), CRM.P2_has_type, DIG.D3_Formal_Derivation))
        new_graph.add((getattr(NGO, derivation_event), DIG.L21_used_as_derivation_source, getattr(NGO, pyramid_PID)))
        new_graph.add((getattr(NGO, derivation_event), DIG.L22_created_derivative, getattr(NGO, thumbnail_PID)))
        new_graph.add((getattr(NGO, thumbnail_PID), RDF.type, CRM.E24_Physical_Human_Made_Thing))
        new_graph.add((getattr(NGO, thumbnail_PID), CRM.P2_has_type, CRM.E24_Physical_Human_Made_Thing))
        new_graph.add((getattr(NGO, thumbnail_PID), CRM.P2_has_type, getattr(WD, 'Q873806')))
        new_graph.add((getattr(WD, 'Q873806'), RDFS.label, Literal('thumbnail', lang="en")))
        new_graph.add((getattr(NGO, thumbnail_PID), CRM.P48_has_preferred_identifier, thumbnail_ID))
        new_graph.add((thumbnail_ID, RDF.value, Literal(obj, lang="en")))
        new_graph.add((thumbnail_ID, RDF.type, CRM.E42_Identifier))
        new_graph.add((thumbnail_ID, CRM.P2_has_type, CRM.E42_Identifier))
        new_graph.add((thumbnail_ID, CRM.P2_has_type, getattr(AAT, '300404629')))
        new_graph.add((getattr(AAT, '300404629'), RDFS.label, Literal('uniform resource identifiers', lang="en")))
        new_graph.add((getattr(NGO, derivation_event), DIG.L23_used_software_or_firmware, getattr(NGO, server_PID)))
        new_graph.add((getattr(NGO, server_PID), RDF.type, DIG.D14_Software))
        new_graph.add((getattr(NGO, server_PID), CRM.P2_has_type, DIG.D14_Software))
        new_graph.add((getattr(NGO, server_PID), CRM.P2_has_type, getattr(AAT, '300266043')))
        new_graph.add((getattr(AAT, '300266043'), RDFS.label, Literal('servers (computer)', lang="en")))
        new_graph.add((getattr(NGO, server_PID), RDFS.label, Literal('https://research.ng-london.org.uk/iiif', lang="en")))
        new_graph.add((server_ID, RDF.type, CRM.E42_Identifier))
        new_graph.add((server_ID, CRM.P2_has_type, CRM.E42_Identifier))

        for work in related_work:
            image_PID = generate_placeholder_PID(work)
            visual_item_PID = create_PID_from_triple('visual item', work)
            new_graph.add((getattr(NGO, thumbnail_PID), CRM.P65_shows_visual_item, getattr(NGO, visual_item_PID)))
            new_graph.add((getattr(NGO, thumbnail_PID), CRM.P62_depicts, getattr(NGO, image_PID)))

    elif pred == getattr(RRO, 'RP5.has_bit_depth'):
        dimension_PID = BNode()
        if obj == getattr(RRI, 'RCL243.8-bit'):
            wikidata_ref = 'Q270159'
            wikidata_desc = '8-bit'
        elif obj == getattr(RRI, 'RCL244.16-bit'):
            wikidata_ref = 'Q194368'
            wikidata_desc = '16-bit architecture'

        new_graph.add((getattr(NGO,subject_PID), CRM.P43_has_dimension, dimension_PID))
        new_graph.add((dimension_PID, RDF.type, CRM.E54_Dimension))
        new_graph.add((dimension_PID, CRM.P2_has_type, CRM.E54_Dimension))
        new_graph.add((dimension_PID, CRM.P2_has_type, getattr(WD, wikidata_ref)))
        new_graph.add((getattr(WD, wikidata_ref), RDFS.label, Literal(wikidata_desc)))
        new_graph.add((dimension_PID, CRM.P2_has_type, getattr(AAT, '300410392')))
        new_graph.add((getattr(AAT, '300410392'), RDFS.label, Literal('bit depth', lang="en")))
        new_graph.add((dimension_PID, CRM.P91_has_unit, getattr(AAT, '300265866')))
        new_graph.add((getattr(AAT, '300265866'), RDFS.label, Literal('bits (computing)', lang="en")))

    return new_graph

def create_examination_event_triples(old_graph, subject_PID, subj, pred, obj, doc_type):
    if pred == getattr(RRO, 'RP98.is_in_project_category'):
        if obj == getattr(RRI, 'RCL211.X-Ray_Images') or obj == getattr(RRI, 'RCL210.X-Ray_Examination'):
            technique_name = 'x-ray'
            technique_event_title = 'X-Ray Examination of '
            aat_value = '300419323'
            aat_title = 'x-radiography'
            object_type_ref = '300056058'
            object_type = 'x-ray (radiation)'
        elif obj == getattr(RRI, 'RCL212.UV_Examination') or obj == getattr(RRI, 'RCL213.UV_Images'):
            technique_name = 'UV'
            technique_event_title = 'UV Examination of '
            aat_value = '300053465'
            aat_title = 'ultraviolet photography'
            object_type_ref = '300181267'
            object_type = 'ultraviolet lamps'
        elif obj == getattr(RRI, 'RCL209.Infrared_Reflectography'):
            technique_name = 'infrared reflectography'
            technique_event_title = 'Infrared Reflectography of '
            aat_value = '300379538'
            aat_title = 'infrared reflectography'
            object_type_ref = '300182292'
            object_type = 'infrared lamps'
        elif obj == getattr(RRI, 'RCL208.Infrared_Photography') or obj == getattr(RRI, 'RCL207.Infrared_Examination'):
            technique_name = 'infrared photography'
            technique_event_title = 'Infrared Photography of '
            aat_value = '300053463'
            aat_title = 'infrared photography'
            object_type_ref = '300182292'
            object_type = 'infrared lamps'
        elif obj == getattr(RRI, 'RCL284.Photomicrographs'):
            technique_name = 'photomicrography'
            technique_event_title = 'Photomicrography of '
            aat_value = '300053550'
            aat_title = 'microphotography'
            object_type_ref = '300024594'
            object_type = 'microscopes'
        elif obj == getattr(RRI, 'RCL215.Visible_Light_Images') or obj == getattr(RRI, 'RCL196.Images_of_Frames') or obj == getattr(RRI, 'RCL214.Visible_Light_Examination'):
            technique_name = 'visible light imaging'
            technique_event_title = 'Visible Light Image of '
            aat_value = '300054225'
            aat_title = 'photography (process)'
            object_type_ref = '300022636'
            object_type = 'cameras (photographic equipment)'
        elif obj == getattr(RRI, 'RCL204.Cross_Sections'):
            technique_name = 'cross section sampling'
            technique_event_title = 'Cross Section Sampling of '
            aat_value = '300034254'
            aat_title = 'cross sections'
            object_type_ref = None
        elif obj == getattr(RRI, 'RCL204.Cross_Sections') and "sample" in subj:
            technique_name = 'microscopic examination'
            technique_event_title = 'Microscopic Examination of '
            aat_value = '300054100'
            aat_title = 'microscopy'
            object_type_ref = '300024594'
            object_type = 'microscopes'
        elif obj == getattr(RRI, 'RCL201.Paint_Binding_Medium'):
            technique_name = 'sampling'
            technique_event_title = 'Sampling of '
            aat_value = '300379429'
            aat_title = 'sampling'
            object_type_ref = None
        elif obj == getattr(RRI, 'RCL203.Microscopy'):
            technique_name = 'microscopic examination'
            technique_event_title = 'Microscopic Examination of '
            aat_value = '300054100'
            aat_title = 'microscopy'
            object_type_ref = '300024594'
            object_type = 'microscopes'
        elif obj == getattr(RRI, 'RCL267.SEM_Examination'):
            technique_name = 'SEM examination'
            technique_event_title = 'SEM Examination of '
            aat_value = '300224955'
            aat_title = 'electron microscopy'
            object_type_ref = '300343819'
            object_type = 'scanning electron microscopes'
        else:
            return new_graph
        
        related_works = query_objects(old_graph, subj, getattr(RRO, 'RP40.is_related_to'), None)
        for work in related_works:
            related_painting_PID = generate_placeholder_PID(work)
            xray_event = create_PID_from_triple(technique_name, work)
            xray_event_title = technique_event_title + work
            
            new_graph.add((getattr(NGO, xray_event), RDFS.label, Literal(xray_event_title, lang="en")))
            new_graph.add((getattr(NGO, xray_event), RDF.type, CRM.E16_Measurement))
            new_graph.add((getattr(NGO, xray_event), CRM.P2_has_type, CRM.E16_Measurement))
            new_graph.add((getattr(NGO, xray_event), CRM.P39_measured, getattr(NGO, related_painting_PID)))
            new_graph.add((getattr(NGO, xray_event), CRM.P32_used_general_technique, getattr(AAT, aat_value)))
            new_graph.add((getattr(AAT, aat_value), RDFS.label, Literal(aat_title, lang="en")))
            new_graph.add((getattr(NGO, xray_event), CRM.P2_has_type, getattr(AAT, '300379861')))
            new_graph.add((getattr(AAT, '300379861'), RDFS.label, Literal('imaging (image making process)', lang="en")))

            if doc_type == 'image':
                new_graph.add((getattr(NGO, subject_PID), CRM.P108i_was_produced_by, getattr(NGO, xray_event)))
            else:
                new_graph.add((getattr(NGO, subject_PID), CRM.P67_refers_to, getattr(NGO, xray_event)))
                new_graph.add((getattr(NGO, subject_PID), CRM.P67_refers_to, getattr(NGO, related_painting_PID)))

            if object_type_ref is not None:
                new_graph.add((getattr(NGO, xray_event), CRM.P125_used_object_of_type, getattr(AAT, object_type_ref)))
                new_graph.add((getattr(AAT, object_type_ref), RDFS.label, Literal(object_type, lang="en")))

        return new_graph

def create_modification_event_triples(old_graph, subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP98.is_in_project_category'):
        if obj == getattr(RRI, 'RCL192.Conservation'):
            related_works = query_objects(old_graph, subj, getattr(RRO, 'RP40.is_related_to'), None)
            for work in related_works:
                related_painting_PID = generate_placeholder_PID(work)
                conservation_event = create_PID_from_triple('conservation', work)
                conservation_event_title = 'Conservation of ' + work

                new_graph.add((getattr(NGO, subject_PID), CRM.P70_documents, getattr(NGO, conservation_event)))
                new_graph.add((getattr(NGO, conservation_event), RDFS.label, Literal(conservation_event_title, lang="en")))
                new_graph.add((getattr(NGO, conservation_event), RDF.type, CRM.E11_Modification))
                new_graph.add((getattr(NGO, conservation_event), CRM.P2_has_type, CRM.E11_Modification))
                new_graph.add((getattr(NGO, conservation_event), CRM.P31_has_modified, getattr(NGO, related_painting_PID)))
                new_graph.add((getattr(NGO, conservation_event), CRM.P2_has_type, getattr(AAT, '300404519')))
                new_graph.add((getattr(AAT, '300404519'), RDFS.label, Literal('conservation (process)', lang="en")))
                new_graph.add((getattr(NGO, subject_PID), CRM.P67_refers_to, getattr(NGO, related_painting_PID)))

    return new_graph   

def create_image_production_event_triples(old_graph, subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP98.is_in_project_category'):
        if obj == getattr(RRI, 'RCL187.Drawings'):
            aat_value = '300054196'
            aat_title = 'drawing (image-making)'
        elif obj == getattr(RRI, 'RCL188.Copies'): 
            aat_value = '300015640'
            aat_title = 'copies (derivative objects)'
        elif obj == getattr(RRI, 'RCL189.Prints'):
            aat_value = '300041273'
            aat_title = 'prints (visual works)'
        elif obj == getattr(RRI, 'RCL271.Study_Images'):
            aat_value = '300081053'
            aat_title = 'studies (visual works)'
        elif obj == getattr(RRI, 'RCL197.Frame_Archive') or obj == getattr(RRI, 'RCL195.Framing'):
            aat_value = '300240903'
            aat_title = 'framing (processes)'
        else:
            return new_graph

        related_works = query_objects(old_graph, subj, getattr(RRO, 'RP40.is_related_to'), None)
        creation_event_PID = create_PID_from_triple('creation', obj)

        new_graph.add((getattr(NGO, subject_PID), RDF.type, CRM.E24_Physical_Human_Made_Thing))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, CRM.E24_Physical_Human_Made_Thing))
        new_graph.add((getattr(NGO, subject_PID), CRM.P108i_was_produced_by, getattr(NGO, creation_event_PID)))
        new_graph.add((getattr(NGO, creation_event_PID), RDF.type, CRM.E12_Production))
        new_graph.add((getattr(NGO, creation_event_PID), CRM.P2_has_type, CRM.E12_Production))
        new_graph.add((getattr(NGO, creation_event_PID), CRM.P2_has_type, getattr(AAT, '300404387')))
        new_graph.add((getattr(AAT, '300404387'), RDFS.label, Literal('creating (artistic activity)', lang="en")))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT, aat_value)))
        new_graph.add((getattr(AAT, aat_value), RDFS.label, Literal(aat_title, lang="en")))

        for work in related_works:
            related_painting_PID = generate_placeholder_PID(work)
            new_graph.add((getattr(NGO, subject_PID), CRM.P62_depicts, getattr(NGO, related_painting_PID)))

    return new_graph

def create_provenance_triples(old_graph, subject_PID, subj, pred, obj):
    if pred == getattr(RRO, 'RP98.is_in_project_category'):
        if obj == getattr(RRI, 'RCL183.Provenance'):
            related_works = query_objects(old_graph, subj, getattr(RRO, 'RP40.is_related_to'), None)
            document_PID = create_PID_from_triple('text', subj)

            for work in related_works:
                related_painting_PID = generate_placeholder_PID(work)
                related_painting_history_event_PID = create_PID_from_triple('provenance', work)
                related_painting_history_event = 'Provenance of ' + work

                new_graph.add((getattr(NGO, subject_PID), CRM.P138_represents, getattr(NGO, document_PID)))
                new_graph.add((getattr(NGO, document_PID), RDF.type, CRM.E31_Document))
                new_graph.add((getattr(NGO, document_PID), CRM.P2_has_type, CRM.E31_Document))
                new_graph.add((getattr(NGO, document_PID), CRM.P70_documents, getattr(RRO, related_painting_history_event_PID)))
                new_graph.add((getattr(NGO, document_PID), CRM.P67_refers_to, getattr(NGO, related_painting_PID)))
                new_graph.add((getattr(NGO, related_painting_history_event_PID), RDF.type, CRM.E5_Event))
                new_graph.add((getattr(NGO, related_painting_history_event_PID), CRM.P2_has_type, CRM.E5_Event))
                new_graph.add((getattr(NGO, related_painting_history_event_PID), CRM.P2_has_type, getattr(AAT, 'th')))
                new_graph.add((getattr(AAT, '300055863'), RDFS.label, Literal('provenance (history of ownership)', lang="en")))
                new_graph.add((getattr(NGO, related_painting_history_event_PID), RDFS.label, Literal(related_painting_history_event, lang="en")))
                new_graph.add((getattr(NGO, related_painting_history_event_PID), CRM.P12_occurred_in_the_presence_of, getattr(NGO, related_painting_PID)))

    return new_graph

def create_sampling_triples(old_graph, subject_PID, subj, pred, obj):
    if obj == getattr(RRO, 'RC23.Sample') or obj == getattr(RRI, 'RCL266.Unmounted_Samples'):
        sampling_event = BNode()
        sampled_section = BNode()
        sampled_works = query_objects(old_graph, subj, getattr(RRO, 'RP52.was_part_of'), None)
        ss_labels = query_objects(old_graph, subj, getattr(RRO, 'RP59.has_description'), None)
        if len(ss_labels) > 0:
            sample_site_label = ss_labels[0]
            new_graph.add((sampled_section, RDFS.label, Literal(sample_site_label, lang="en")))
        national_gallery_staff_member_PID = generate_placeholder_PID('National Gallery staff member')
        doc_BN = BNode()
        doc2_BN = BNode()
        alteration_BN = BNode()

        new_graph.add((sampling_event, RDF.type, SCI.S2_Sample_Taking))
        new_graph.add((sampling_event, CRM.P2_has_type, SCI.S2_Sample_Taking))
        new_graph.add((sampling_event, SCI.O5_removed, getattr(NGO, subject_PID)))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT, '300028875')))
        new_graph.add((getattr(AAT, '300028875'), RDFS.label, Literal('samples', lang="en")))
        new_graph.add((getattr(NGO, subject_PID), RDF.type, SCI.S13_Sample))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, SCI.S13_Sample))
        new_graph.add((sampling_event, CRM.P14_carried_out_by, getattr(NGO, national_gallery_staff_member_PID)))
        new_graph.add((getattr(NGO, national_gallery_staff_member_PID), RDF.type, CRM.E21_Person))
        new_graph.add((getattr(NGO, national_gallery_staff_member_PID), CRM.P2_has_type, CRM.E21_Person))
        new_graph.add((getattr(NGO, national_gallery_staff_member_PID), CRM.P2_has_type, getattr(AAT, '300025820')))
        new_graph.add((getattr(AAT, '300025820'), RDFS.label, Literal('conservation scientist', lang="en")))
        new_graph.add((getattr(NGO, national_gallery_staff_member_PID), RDFS.label, Literal('Example NG Staff Member', lang="en")))
        new_graph.add((sampled_section, RDF.type, CRM.E53_Place))
        new_graph.add((sampled_section, CRM.P2_has_type, CRM.E53_Place))
        new_graph.add((sampled_section, CRM.P2_has_type, getattr(AAT, '300241583')))
        new_graph.add((getattr(AAT, '300241583'), RDFS.label, Literal('components (object part)', lang="en")))
        new_graph.add((sampling_event, SCI.O4_sampled_at, sampled_section))
        new_graph.add((sampling_event, CRM.P70_is_documented_in, doc_BN))
        new_graph.add((doc_BN, RDF.type, CRM.E31_Document))
        new_graph.add((doc_BN, CRM.P2_has_type, CRM.E31_Document))
        new_graph.add((doc_BN, CRM.P2_has_type, getattr(WD, 'Q1710397')))
        new_graph.add((getattr(WD, 'Q1710397'), RDFS.label, Literal('reason', lang="en")))
        new_graph.add((doc_BN, RDFS.comment, Literal('As part of a general technical examination of the painting - documenting the material and techniques used in its production.', lang="en")))
        new_graph.add((sampling_event, RDFS.comment, Literal('Actual textual content describing the sampling process', lang="en")))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, getattr(AAT, '300034254')))
        new_graph.add((getattr(NGO, subject_PID), RDFS.label, Literal(subj, lang="en")))

        for work in sampled_works:
            sampled_work_PID = generate_placeholder_PID(work)

            new_graph.add((getattr(NGO, sampled_work_PID), CRM.P59_has_section, sampled_section))
            new_graph.add((sampling_event, SCI.O3_sampled_from, getattr(NGO, sampled_work_PID)))
            new_graph.add((getattr(NGO, subject_PID), SCI.O25i_is_contained_in, getattr(NGO, sampled_work_PID)))

        if obj == getattr(RRO, 'RC23.Sample'):
            new_graph.add((getattr(NGO, subject_PID), SCI.O18i_was_altered_by, alteration_BN))
            new_graph.add((alteration_BN, RDF.type, SCI.S18_Alteration))
            new_graph.add((alteration_BN, CRM.P2_has_type, SCI.S18_Alteration))
            new_graph.add((alteration_BN, CRM.P2_has_type, getattr(AAT, '300443464')))
            new_graph.add((getattr(AAT, '300443464'), RDFS.label, Literal('embedding', lang="en")))
            new_graph.add((alteration_BN, CRM.P2_has_type, CRM.E13_Attribute_Assignment))
            new_graph.add((alteration_BN, CRM.P140_assigned_attribute_to, getattr(NGO, subject_PID)))
            new_graph.add((alteration_BN, CRM.P177_assigned_property_of_type, getattr(AAT, '300034254')))
            new_graph.add((getattr(AAT, '300034254'), RDFS.label, Literal('cross sections', lang="en")))
            new_graph.add((alteration_BN, CRM.P16_used_specific_object, getattr(AAT, '300014533')))
            new_graph.add((getattr(AAT, '300014533'), RDFS.label, Literal('epoxy resin', lang="en")))

    return new_graph        
    
def map_object(old_graph, new_graph):
    for painting_id, _, _ in old_graph.triples((None, RDF.type, getattr(RRO,'RC12.Painting'))):
        for subj, pred, obj in old_graph.triples((painting_id, None, None)):
            subject_PID = generate_placeholder_PID(subj)
            new_graph = create_title_triples(subject_PID, subj, pred, obj)
            new_graph = create_medium_triples(subject_PID, subj, pred, obj)
            new_graph = create_collection_triples(subject_PID, subj, pred, obj)
            new_graph = create_dimension_triples(subject_PID, subj, pred, obj)
            new_graph = create_identifier_triples(subject_PID, pred, obj)
            new_graph = create_type_triples(subject_PID, pred, obj)
            new_graph = create_location_triples(subject_PID, subj, pred, obj)

    return new_graph

def map_event(old_graph, new_graph):
    event_properties_list = [
        'RP72.was_produced',
        'RP68.was_acquired',
        'RP42.was_born_in',
        'RP4.died_in'
    ]

    for event_property in event_properties_list:
        for entity, _, event_name in old_graph.triples((None, getattr(RRO, event_property), None)):
            for subj, pred, obj in old_graph.triples((entity, None, event_name)):
                subject_PID = generate_placeholder_PID(subj)
                event_PID = generate_placeholder_PID(obj)
                new_graph = create_event_triples(subject_PID, event_PID, subj, pred)
            for subj, pred, obj in old_graph.triples((event_name, None, None)):
                subject_PID = generate_placeholder_PID(subj)
                new_graph = create_time_span_triples(subject_PID, subj, pred, obj)
                new_graph = create_comment_triples(subject_PID, pred, obj)
                new_graph = create_actor_event_relationship_triples(subject_PID, pred, obj)
                #location

    return new_graph

def map_person(old_graph, new_graph):
    for person_name, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC40.Person'))):
        for subj, pred, obj in old_graph.triples((person_name, None, None)):
            subject_PID = generate_placeholder_PID(subj)
            new_graph = create_type_triples(subject_PID, pred, obj)
            new_graph = create_name_triples(subject_PID, subj, pred, obj)
            new_graph = create_role_triples(subject_PID, pred, obj)
            new_graph = create_comment_triples(subject_PID, pred, obj)
            new_graph = create_location_triples(subject_PID, subj, pred, obj)

    return new_graph

def map_institution(old_graph, new_graph):
    for institution_name, _, _ in old_graph.triples((getattr(RRI, 'The_National_Gallery'), RDF.type, getattr(RRO, 'RC41.Institution'))):
        subject_PID = generate_placeholder_PID(institution_name)
        
        wikidata_name = wikidata_query(institution_name, 'institution')
        if wikidata_name is not None:
            new_graph.add((getattr(NGO, subject_PID), OWL.sameAs, getattr(WD, wikidata_name)))
        
        new_graph = create_institution_triples(old_graph, new_graph, subject_PID, institution_name)
    
    for room_name, _, _ in old_graph.triples((getattr(RRI, 'Room_8'), RDF.type, getattr(RRO, 'RC11.Room'))):
        subject_PID = generate_placeholder_PID(room_name)

        for subj, pred, obj in old_graph.triples((room_name, None, None)):
            new_graph = create_room_triples(subject_PID, subj, pred, obj)

    for area_name, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC264.Area_in_Room'))):
        subject_PID = generate_placeholder_PID(area_name)

        for subj, pred, obj in old_graph.triples((area_name, None, None)):
            new_graph = create_area_of_room_triples(subject_PID, subj, pred, obj)
    
    return new_graph

def map_document(old_graph, new_graph):
    
    for doc_name, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC26.Digital_Document'))):
        subject_PID = generate_placeholder_PID(doc_name)
        database_PID = generate_placeholder_PID('The National Gallery Collection Image Database')
        text_PID = create_PID_from_triple('text', doc_name)

        new_graph.add((getattr(NGO, subject_PID), RDF.type, DIG.D1_Digital_Object))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, DIG.D1_Digital_Object))
        new_graph.add((getattr(NGO, subject_PID), CRM.P70i_is_documented_in, getattr(NGO, database_PID)))
        new_graph.add((getattr(NGO, database_PID), RDFS.label, Literal('The National Gallery Collection Image Database', lang="en")))
        new_graph.add((getattr(NGO, subject_PID), CRM.P165_incorporates, getattr(NGO, text_PID)))
        new_graph.add((getattr(NGO, text_PID), RDF.type, CRM.E90_Symbolic_Object))
        new_graph.add((getattr(NGO, text_PID), CRM.P2_has_type, CRM.E90_Symbolic_Object))

        for subj, pred, obj in old_graph.triples((doc_name, None, None)):
            new_graph = create_type_triples(subject_PID, pred, obj)
            new_graph = create_reference_triples(subject_PID, subj, pred, obj)
            new_graph = create_file_triples(g, subject_PID, subj, pred, obj)
            new_graph = create_examination_event_triples(g, subject_PID, subj, pred, obj, doc_type='document')
            new_graph = create_actor_event_relationship_triples(subject_PID, pred, obj)
    
    for doc_name, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC220.Digital_Text'))):
        subject_PID = generate_placeholder_PID(doc_name)
        database_PID = generate_placeholder_PID('The Raohael Research Resource')

        new_graph.add((getattr(NGO, subject_PID), CRM.P70i_is_documented_in, getattr(NGO, database_PID)))
        new_graph.add((getattr(NGO, database_PID), RDFS.label, Literal('The Raphael Research Resource', lang="en")))

        for subj, pred, obj in old_graph.triples((doc_name, None, None)):    
            new_graph = create_type_triples(subject_PID, pred, obj)
            new_graph = create_reference_triples(subject_PID, subj, pred, obj)
            new_graph = create_file_triples(g, subject_PID, subj, pred, obj)
            new_graph = create_examination_event_triples(g, subject_PID, subj, pred, obj, doc_type='document')
            new_graph = create_actor_event_relationship_triples(subject_PID, pred, obj)
            new_graph = create_comment_triples(subject_PID, pred, obj)

            if pred == getattr(RRO, 'RP98.is_in_project_category') and (obj == getattr(RRI, 'RCL184.General_Bibliography') or obj == getattr(RRI, 'RCL185.Exhibition_and_Loan_History')):
                related_bibliography = query_objects(old_graph, subj, getattr(RRO, 'RP237.has_content'), None)
                painting_id = query_objects(old_graph, subj, getattr(RRO, 'RP40.is_related_to'), None)[0]
                for _,_,o in old_graph.triples((subj, getattr(RRO, 'RP237.has_content'), None)):
                    references_list = o.split('\n')
                    create_triples_from_reference_string(references_list, painting_id)
    
    return new_graph

def map_image(old_graph, new_graph):
    for image_name, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC25.Image'))):
        subject_PID = generate_placeholder_PID(image_name)
        related_works = query_objects(old_graph, image_name, getattr(RRO, 'RP40.is_related_to'), None)
        database_PID = generate_placeholder_PID('The National Gallery Collection Image Database')

        new_graph.add((getattr(NGO, subject_PID), RDF.type, DIG.D1_Digital_Object))
        new_graph.add((getattr(NGO, subject_PID), CRM.P2_has_type, DIG.D1_Digital_Object))
        new_graph.add((getattr(NGO, subject_PID), CRM.P70i_is_documented_in, getattr(NGO, database_PID)))
        new_graph.add((getattr(NGO, database_PID), RDFS.label, Literal('The National Gallery Collection Image Database', lang="en")))

        if (image_name, getattr(RRO, 'RP98.is_in_project_category'), getattr(RRI, 'RCL183.Provenance')) in old_graph:
            doc_type = 'document'
        else:
            doc_type = 'image'

        #only for images that are actually pictures of paintings; otherwise they must be treated as documents - how do we define this?????
        for work in related_works:
            if (getattr(RRI, work), RDF.type, getattr(RRO, 'RC12.Painting')) in old_graph:
                painting_PID = generate_placeholder_PID(work)
                image_PID = create_PID_from_triple('visual item', work)

                new_graph.add((getattr(NGO, subject_PID), CRM.P65_shows_visual_item, getattr(NGO, image_PID)))
                new_graph.add((getattr(NGO, image_PID), RDF.type, CRM.E36_Visual_Item))
                new_graph.add((getattr(NGO, image_PID), CRM.P2_has_type, CRM.E36_Visual_Item))
                new_graph.add((getattr(NGO, image_PID), CRM.P1i_is_identified_by, Literal("Visual image of " + work, lang="en")))
                new_graph.add((getattr(NGO, image_PID), CRM.P138_represents, getattr(NGO, painting_PID)))
                new_graph.add((getattr(NGO, subject_PID), CRM.P62_depicts, getattr(NGO, painting_PID)))

        for subj, pred, obj in old_graph.triples((image_name, None, None)):
            new_graph = create_file_triples(g, subject_PID, subj, pred, obj)
            new_graph = create_dimension_triples(subject_PID, subj, pred, obj)
            new_graph = create_type_triples(subject_PID, pred, obj)
            new_graph = create_modification_event_triples(g, subject_PID, subj, pred, obj)
            new_graph = create_examination_event_triples(g, subject_PID, subj, pred, obj, doc_type=doc_type)
            new_graph = create_image_production_event_triples(g, subject_PID, subj, pred, obj)
            new_graph = create_provenance_triples(g, subject_PID, subj, pred, obj)
            new_graph = create_reference_triples(subject_PID, subj, pred, obj)

    return new_graph

def map_sample(old_graph, new_graph):
    for sample_name, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC23.Sample'))):
        subject_PID = generate_placeholder_PID(sample_name)

        for subj, pred, obj in old_graph.triples((sample_name, None, None)):
            new_graph = create_sampling_triples(g, subject_PID, subj, pred, obj)
            new_graph = create_examination_event_triples(g, subject_PID, subj, pred, obj, doc_type='image')

    return new_graph

def map_leftover_categories(old_graph, new_graph):
    
    for file_path, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC223.Computer_Path'))):
        new_graph.add((Literal(file_path), RDF.type, CRM.E73_Information_Object))
        new_graph.add((Literal(file_path), CRM.P2_has_type, CRM.E73_Information_Object))
    for file_path, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC280.IIPImage_Server'))):
        new_graph.add((Literal(file_path), RDF.type, CRM.E73_Information_Object))
        new_graph.add((Literal(file_path), CRM.P2_has_type, CRM.E73_Information_Object))
    for file_path, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC287.Commercial_Link'))):
        new_graph.add((Literal(file_path), RDF.type, CRM.E73_Information_Object))
        new_graph.add((Literal(file_path), CRM.P2_has_type, CRM.E73_Information_Object))
    
    for boolean_obj, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC227.Boolean'))):
        if boolean_obj == getattr(RRI, 'RCL228.Yes'):
            boolean_literal = 'Yes'
        elif boolean_obj == getattr(RRI, 'RCL229.No'):
            boolean_literal = 'No'

        new_graph.add((Literal(boolean_literal), RDF.type, CRM.E59_Primitive_Value))
        new_graph.add((Literal(boolean_literal), CRM.P2_has_type, CRM.E59_Primitive_Value))
    
    for language, _, _ in old_graph.triples((None, RDF.type, getattr(RRO, 'RC232.Language'))):
        language_bn = BNode()
        for subj, pred, obj in old_graph.triples((language, None, None)): 
            if pred == RDF.type:
                new_graph.add((language_bn, RDF.type, CRM.E56_Language))
                new_graph.add((language_bn, CRM.P2_has_type, CRM.E56_Language))
                new_graph.add((language_bn, RDFS.label, Literal(subj, lang="en")))
            elif pred == getattr(RRO, 'RP56.has_name'):
                wd_result = wikidata_query(obj, 'language')
                new_graph.add((language_bn, CRM.P72_has_language, Literal(obj, lang="en")))
                if wd_result is not None:
                    new_graph.add((language_bn, CRM.P72_has_language, getattr(WD, wd_result)))
    
    return new_graph

#def output_example_for_visualisation(new_graph):
#print(sparql_query_pythonic())

#new_graph = map_property(g, RDF.type, RDFS.subClassOf)
#query_graph(new_graph,None,RDFS.subClassOf,None)

#new_dataframe = sparql_query_pythonic('NG1171',csv_format=False)
#mapped_dataframe = map_property(new_dataframe,'NG1171','PID')
#print(mapped_dataframe)
#triples_to_csv(new_dataframe)

#generate_placeholder_PID('Fake Painting 3')

#query_graph(g, None, getattr(RRO,'RP34.has_title'), None)
#map_property(g, new_graph, getattr(RRO,'RP34.has_title'), getattr(CRM,'P102_Has_Title'))
#query_graph(g, getattr(RRI, 'N-0168-00-000107'), getattr(RRO, 'RP15.has_format'), getattr(RRI, 'RCL89.Tiff'))
#triples_to_csv(new_graph,'crm_mapping_draft')
db = connect_to_sql()

map_object(g, new_graph)
map_event(g, new_graph)
map_image(g, new_graph)
map_institution(g, new_graph)
map_person(g, new_graph)
map_document(g, new_graph)
map_sample(g, new_graph)
map_leftover_categories(g, new_graph)
for x,y,z in new_graph:
    print(x.n3(new_graph.namespace_manager), y.n3(new_graph.namespace_manager), z.n3(new_graph.namespace_manager))
#triples_to_tsv(new_graph,'sample_comparison')
new_graph.serialize(destination='outputs/raphael_final.xml', format='xml')
new_graph.serialize(destination='outputs/raphael_final.ttl', format='ttl')
new_graph.serialize(destination='outputs/raphael_final.trig', format='trig')
print('Finished running, all looking good')